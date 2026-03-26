import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# User-Agent 설정 (웹사이트가 크롤러를 차단할 수 있으므로 필요)
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# Naver 검색 URL
url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=%EB%B0%98%EB%8F%84%EC%B2%B4"

try:
    # 웹 페이지 요청
    response = requests.get(url, headers=headers, timeout=10)
    response.encoding = 'utf-8'
    
    # 상태 코드 확인
    if response.status_code == 200:
        # BeautifulSoup으로 파싱
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 신문기사 제목을 포함한 span 태그 찾기
        # sds-comps-text-type-headline1 또는 sds-comps-text-type-body2 클래스의 span 태그
        title_spans = soup.find_all('span', class_=re.compile(r'sds-comps-text-type-(headline|body)'))
        
        titles = []
        
        # 각 span에서 제목 추출 (중복 제거)
        for span in title_spans:
            text = span.get_text(strip=True)
            # <mark> 태그 내용은 제거 (마크된 부분)
            text = re.sub(r'<mark>.*?</mark>', '', text)
            # 공백 제거
            text = text.strip()
            
            # 너무 짧은 텍스트는 제외 (5자 이상만)
            if text and len(text) > 5 and text not in titles:
                titles.append(text)
        
        print(f"총 {len(titles)}개의 뉴스 제목을 찾았습니다.\n")
        print("=" * 100)
        
        # 중복을 제거한 제목 출력
        for i, title in enumerate(titles, 1):
            # 200자 이상인 경우 요약
            if len(title) > 100:
                print(f"[{i}] {title[:100]}...")
            else:
                print(f"[{i}] {title}")
        
        print("=" * 100)
        
        # Excel 파일로 저장
        try:
            # 새로운 Workbook 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "뉴스기사"
            
            # 헤더 설정
            headers_row = ["번호", "기사 제목", "크롤링 일시"]
            ws.append(headers_row)
            
            # 헤더 스타일 설정 (파란색 배경, 흰색 글자, 굵게)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터 추가
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for i, title in enumerate(titles, 1):
                ws.append([i, title, current_time])
            
            # 열 너비 자동 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 80
            ws.column_dimensions['C'].width = 20
            
            # 모든 셀에 텍스트 줄바꿈 활성화
            for row in ws.iter_rows(min_row=2, max_row=len(titles)+1, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # 파일 저장
            file_path = "naver_result.xlsx"
            wb.save(file_path)
            print(f"\n✓ 크롤링 결과가 '{file_path}' 파일로 저장되었습니다.")
            
        except Exception as e:
            print(f"\n✗ Excel 파일 저장 중 오류: {e}")
    else:
        print(f"요청 실패: 상태 코드 {response.status_code}")
        
except requests.exceptions.RequestException as e:
    print(f"요청 중 오류 발생: {e}")
except Exception as e:
    print(f"오류 발생: {e}")
