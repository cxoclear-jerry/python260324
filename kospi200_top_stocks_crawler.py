import sys
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from PyQt6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QTextEdit, QProgressBar, QLabel, QMessageBox
from PyQt6.QtCore import QThread, pyqtSignal
from openpyxl import Workbook

class CrawlerThread(QThread):
    progress = pyqtSignal(int)  # 진행률
    log = pyqtSignal(str)       # 로그 메시지
    finished_signal = pyqtSignal(list)  # 완료된 데이터

    def run(self):
        base_url = "https://finance.naver.com/sise/entryJongmok.naver?type=KPI200&page="
        all_stocks = []
        
        for page in range(1, 21):
            self.log.emit(f"페이지 {page} 크롤링 중...")
            
            try:
                response = requests.get(f"{base_url}{page}")
                response.raise_for_status()
                
                soup = BeautifulSoup(response.text, 'html.parser')
                table = soup.find('table', {'class': 'type_1'})
                
                if not table:
                    self.log.emit(f"페이지 {page}: 테이블을 찾을 수 없습니다.")
                    continue
                
                rows = table.find_all('tr')
                page_stocks = []
                
                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) == 7 and cols[0].get('class') == ['ctg']:
                        stock_data = {
                            '종목명': cols[0].text.strip(),
                            '현재가': cols[1].text.strip(),
                            '전일비': cols[2].text.strip(),
                            '등락률': cols[3].text.strip(),
                            '거래량': cols[4].text.strip(),
                            '거래대금(백만)': cols[5].text.strip(),
                            '시가총액(억)': cols[6].text.strip()
                        }
                        page_stocks.append(stock_data)
                
                if not page_stocks:
                    self.log.emit(f"페이지 {page}: 데이터가 없습니다. 크롤링 종료.")
                    break
                
                all_stocks.extend(page_stocks)
                self.log.emit(f"페이지 {page}: {len(page_stocks)}개 종목 수집")
                self.progress.emit(page * 5)  # 20페이지, 5%씩
                
                time.sleep(1)
                
            except Exception as e:
                self.log.emit(f"페이지 {page} 오류: {e}")
                continue
        
        self.progress.emit(100)
        self.finished_signal.emit(all_stocks)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("코스피200 편입종목 크롤러")
        self.setGeometry(100, 100, 800, 600)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        # 버튼 레이아웃
        button_layout = QHBoxLayout()
        self.start_button = QPushButton("크롤링 시작")
        self.start_button.clicked.connect(self.start_crawling)
        button_layout.addWidget(self.start_button)
        
        self.save_button = QPushButton("엑셀로 저장")
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setEnabled(False)  # 초기에는 비활성화
        button_layout.addWidget(self.save_button)
        
        layout.addLayout(button_layout)
        
        # 프로그레스 바
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)
        
        # 로그 텍스트
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)
        
        # 결과 라벨
        self.result_label = QLabel("결과가 여기에 표시됩니다.")
        layout.addWidget(self.result_label)
        
        self.crawler_thread = None
        self.stocks = []  # 크롤링 데이터 저장
    
    def start_crawling(self):
        self.start_button.setEnabled(False)
        self.save_button.setEnabled(False)
        self.log_text.clear()
        self.progress_bar.setValue(0)
        self.result_label.setText("크롤링 중...")
        
        self.crawler_thread = CrawlerThread()
        self.crawler_thread.progress.connect(self.progress_bar.setValue)
        self.crawler_thread.log.connect(self.log_text.append)
        self.crawler_thread.finished_signal.connect(self.on_finished)
        self.crawler_thread.start()
    
    def on_finished(self, stocks):
        self.start_button.setEnabled(True)
        self.save_button.setEnabled(True)
        self.stocks = stocks
        self.result_label.setText(f"총 {len(stocks)}개 종목 수집 완료")
        
        # 결과를 로그에 추가
        self.log_text.append(f"\n총 {len(stocks)}개 종목 수집 완료:")
        self.log_text.append("-" * 80)
        for i, stock in enumerate(stocks, 1):
            self.log_text.append(f"{i}. 종목명: {stock['종목명']}")
            self.log_text.append(f"   현재가: {stock['현재가']}")
            self.log_text.append(f"   전일비: {stock['전일비']}")
            self.log_text.append(f"   등락률: {stock['등락률']}")
            self.log_text.append(f"   거래량: {stock['거래량']}")
            self.log_text.append(f"   거래대금(백만): {stock['거래대금(백만)']}")
            self.log_text.append(f"   시가총액(억): {stock['시가총액(억)']}")
            self.log_text.append("-" * 40)
    
    def save_to_excel(self):
        if not self.stocks:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 크롤링을 실행하세요.")
            return
        
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"kospi200[{today}].xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "코스피200 편입종목"
            
            # 헤더
            headers = ["종목명", "현재가", "전일비", "등락률", "거래량", "거래대금(백만)", "시가총액(억)"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 데이터
            for row, stock in enumerate(self.stocks, 2):
                ws.cell(row=row, column=1, value=stock['종목명'])
                ws.cell(row=row, column=2, value=stock['현재가'])
                ws.cell(row=row, column=3, value=stock['전일비'])
                ws.cell(row=row, column=4, value=stock['등락률'])
                ws.cell(row=row, column=5, value=stock['거래량'])
                ws.cell(row=row, column=6, value=stock['거래대금(백만)'])
                ws.cell(row=row, column=7, value=stock['시가총액(억)'])
            
            wb.save(filename)
            QMessageBox.information(self, "성공", f"파일이 {filename}으로 저장되었습니다.")
            self.log_text.append(f"엑셀 파일 저장: {filename}")
        
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {e}")
            self.log_text.append(f"저장 오류: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())