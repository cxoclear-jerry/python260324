from openpyxl import Workbook
import random

# 전자제품 목록
products = [
    "삼성 갤럭시 S24", "아이폰 15", "갤럭시 Z 폴드5", "아이폰 15 Pro",
    "삼성 갤럭시 탭 S9", "아이패드 프로", "삼성 갤럭시 북3", "맥북 에어",
    "삼성 갤럭시 워치6", "애플 워치 울트라", "에어팟 프로", "갤럭시 버즈2 프로",
    "아이폰 SE", "갤럭시 A54", "아이패드 에어", "삼성 갤럭시 탭 A8",
    "LG 그램 17", "맥북 프로 14인치", "삼성 갤럭시 북3 프로", "아이패드 미니",
    "갤럭시 워치5", "애플 워치 SE", "소니 WH-1000XM5", "보스 QC 울트라 헤드폰",
    "삼성 갤럭시 버즈 FE", "JBL GO 3", "아이폰 14", "갤럭시 S23",
    "아이폰 15 Pro Max", "갤럭시 Z 플립5", "아이패드", "삼성 갤럭시 탭 S8",
    "레노버 요가 북", "맥북 프로 16인치", "삼성 갤럭시 북2", "아이패드 프로 11인치",
    "갤럭시 워치4", "애플 워치 시리즈8", "소니 WF-1000XM4", "보스 QC 헤드폰",
    "삼성 갤럭시 버즈 프로", "JBL T600", "아이폰 13", "갤럭시 S22",
    "아이폰 14 Pro", "갤럭시 Z 폴드4", "아이패드 10세대", "삼성 갤럭시 탭 S7",
    "에이서 스위프트 5", "맥북 에어 M2", "삼성 갤럭시 북 프로", "아이패드 프로 12.9인치"
]

def generate_product_data():
    """전자제품 판매 데이터를 생성하는 함수"""
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "전자제품 판매 데이터"

    # 헤더 추가
    headers = ["제품ID", "제품명", "수량", "가격"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # 100개의 데이터 생성
    for row_num in range(2, 102):  # 2부터 101까지 (헤더 제외)
        # 제품ID 생성 (P001, P002, ...)
        product_id = f"P{row_num-1:03d}"

        # 랜덤 제품명 선택
        product_name = random.choice(products)

        # 랜덤 수량 생성 (1-100개)
        quantity = random.randint(1, 100)

        # 랜덤 가격 생성 (10,000원 - 3,000,000원)
        price = random.randint(10000, 3000000)

        # 데이터 입력
        ws.cell(row=row_num, column=1, value=product_id)
        ws.cell(row=row_num, column=2, value=product_name)
        ws.cell(row=row_num, column=3, value=quantity)
        ws.cell(row=row_num, column=4, value=price)

    # 파일 저장
    wb.save("products.xlsx")
    print("products.xlsx 파일이 성공적으로 생성되었습니다!")
    print("총 100개의 전자제품 판매 데이터가 저장되었습니다.")

if __name__ == "__main__":
    generate_product_data()